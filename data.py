import urllib.request
import json
import requests
import openpyxl
import random
import re
import os

locations = set()

currentPath = os.getcwd()

os.chdir(currentPath)

def loadLocation():  # 엑셀 파일에 저장된 지역, 대학이름 불러오기
    global locations
    excel = openpyxl.load_workbook("C://Users/user/PycharmProjects/tastebot/static/location.xlsx").worksheets[0]
    # 데이터로 뽑아내기 힘든 지역들 디폴트로 집어넣기
    location = set(
        ['경기', '강원', '전남', '전북', '경북', '경남', '충북', '충남', '서울', '인천', '대구', '부산', '대전', '울산', '광주', '고려대', '홍익대', '명지대',
         '중앙대', '연세대'])
    for row in excel.rows:
        if (row[1].value):  # 시 / 군
            location.add(row[1].value[:-1])
            location.add(row[1].value)
        if (row[2].value):  # 구
            location.add(row[2].value[:-1])
            location.add(row[2].value)
        if (row[3].value):  # 동 / 읍 / 면 / 리
            location.add(row[3].value)

    excel = openpyxl.load_workbook("C://Users/user/PycharmProjects/tastebot/static/school.xlsx").worksheets[0]  # 대학 불러오기
    for row in excel.rows:
        if (row[0].value):  # 'OO대' 로 저장
            if (row[0].value[-2:] == "학교"):
                location.add(row[0].value[:-2])
                location.add(row[0].value)
            elif (row[0].value[-2:] == "대학"):
                location.add(row[0].value[:-2])
                location.add(row[0].value)

    location.remove("")
    locations = list(location)
    locations.sort(key=lambda x: len(x), reverse=True)  # 단어의 길이가 더 긴것부터 검색되도록


foods = set()


def loadFood():
    global foods
    excel = openpyxl.load_workbook("C://Users/user/PycharmProjects/tastebot/static/food.xlsx").worksheets[0]  # 음식 불러오기
    food = set()
    for row in excel.rows:
        if not row[0].value:
            continue
        val = str(row[0].value)
        if (val[-1] != ')' and val[-1] != 'g'):  # 쓸데없는 데이터 제외
            food.add(val)

    foods = list(food)
    foods.sort(key=lambda x: len(x), reverse=True)  # 단어의 길이가 더 긴것부터 검색되도록


loadLocation()
loadFood()


def getLocations():
    return locations


def getFoods():
    return foods


ncreds = {
    "client_id": "YZHKw9SAeIBdqdCCgxQE",
    "client_secret": "eb0ZvmkJbA"
}
nheaders = {
    "X-Naver-Client-Id": ncreds.get('client_id'),
    "X-Naver-Client-Secret": ncreds.get('client_secret')
}

# 네이버 지역 검색 주소
naver_local_url = "https://openapi.naver.com/v1/papago/n2mt"

# 검색에 사용될 파라미터
# 정렬 sort : 리뷰순(comment)
# 검색어 query : 인코딩된 문자열
params_format = "source=ko&target=en&text="


# 위치는 사용자가 사용할 지역으로 변경가능

def searchRestaurant(location="", food="", user_input=""):
    # input으로 위치, 음식, 맛집 특징 user_input을 주고, 맛집의 이름, 카테고리, 주소를 출력해주는 파이썬함수

    name = ""
    category = ""
    address = ""
    print("searchRestaurant에서 받은 user_input값"+user_input)



    info = [name, category, address]
    return info